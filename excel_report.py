"""
excel_report.py
---------------
Tanggung jawab: Menerima PIL Image dari memori, membuat thumbnail,
dan menyisipkannya ke file Excel dengan tampilan profesional.

DESAIN VISUAL:
- Header   : gradient biru gelap (#0D1B2A) dengan font putih bold
- Baris    : warna bergantian (stripe) putih dan biru sangat muda
- Thumbnail: diberi border dan shadow efek
- Sheet    : tab berwarna, title report, freeze panes, gridlines off
- Kolom No : center, background lebih gelap sebagai aksen
- Border   : thin border di semua sel data

ATURAN TEKNIS:
- File temp hanya dihapus di finalize() — openpyxl butuh file temp
  tetap ada sampai save() terakhir selesai.
"""

import tempfile
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill,
    PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage

from config import (
    REPORTS_DIR,
    EXCEL_COLUMNS,
    THUMBNAIL_WIDTH,
    THUMBNAIL_HEIGHT_RATIO,
    EXCEL_ROW_HEIGHT,
    get_excel_filename,
)
from logger import get_logger

logger = get_logger()


# =============================================================================
# PALET WARNA
# =============================================================================
# Semua warna di satu tempat — mudah diubah

COLOR_HEADER_BG      = "0D1B2A"   # Biru sangat gelap (header)
COLOR_HEADER_FONT    = "FFFFFF"   # Putih
COLOR_ACCENT         = "1A73E8"   # Biru Google (aksen)
COLOR_ROW_ODD        = "FFFFFF"   # Putih
COLOR_ROW_EVEN       = "EBF3FD"   # Biru sangat muda
COLOR_NO_CELL        = "1E3A5F"   # Biru gelap untuk kolom No
COLOR_NO_FONT        = "FFFFFF"   # Putih
COLOR_TITLE_BG       = "0D1B2A"   # Sama dengan header
COLOR_TITLE_ROW      = "E8F0FE"   # Biru muda untuk baris info
COLOR_BORDER         = "BDD7EE"   # Border biru muda


class ExcelReportManager:
    """
    Mengelola pembuatan dan pembaruan file Excel report harian
    dengan tampilan profesional.

    Attributes:
        _filepath    : Path file Excel hari ini.
        _workbook    : Objek Workbook openpyxl.
        _worksheet   : Sheet aktif.
        _row_counter : Nomor baris berikutnya.
        _entry_number: Nomor urut entry.
        _temp_files  : List file thumbnail sementara — hapus di finalize().
    """

    def __init__(self) -> None:
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        self._filepath: Path = REPORTS_DIR / get_excel_filename()
        self._temp_files: List[Path] = []

        if self._filepath.exists():
            try:
                self._workbook = load_workbook(self._filepath)
                self._worksheet = self._workbook.active
                self._row_counter = self._worksheet.max_row
                self._entry_number = max(0, self._row_counter - 1)
                logger.info(
                    f"Excel report opened: {self._filepath.name} "
                    f"| Existing entries: {self._entry_number}"
                )
            except Exception as e:
                logger.error(f"Gagal membuka Excel: {e} — membuat file baru.")
                self._create_new()
        else:
            self._create_new()

    # =========================================================================
    # PUBLIC
    # =========================================================================

    def add_entry(self, image: PilImage.Image) -> None:
        """
        Tambahkan satu baris ke Excel dengan thumbnail screenshot.

        Args:
            image: PIL Image dari memori (hasil capture_screenshot).
        """
        self._row_counter += 1
        self._entry_number += 1
        row = self._row_counter
        now = datetime.now()

        # Warna baris bergantian (stripe effect)
        row_color = COLOR_ROW_ODD if self._entry_number % 2 != 0 else COLOR_ROW_EVEN
        row_fill = PatternFill(fill_type="solid", fgColor=row_color)

        # Thin border untuk semua sel
        thin = Side(border_style="thin", color=COLOR_BORDER)
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Kolom A: No ───────────────────────────────────────────────────
        cell_no = self._worksheet.cell(row=row, column=1)
        cell_no.value = self._entry_number
        cell_no.fill = PatternFill(fill_type="solid", fgColor=COLOR_NO_CELL)
        cell_no.font = Font(color=COLOR_NO_FONT, bold=True, size=11)
        cell_no.alignment = Alignment(horizontal="center", vertical="center")
        cell_no.border = cell_border

        # ── Kolom B: Tanggal ──────────────────────────────────────────────
        cell_date = self._worksheet.cell(row=row, column=2)
        cell_date.value = now.strftime("%d-%m-%Y")
        cell_date.fill = row_fill
        cell_date.font = Font(size=11, color="2C3E50")
        cell_date.alignment = Alignment(horizontal="center", vertical="center")
        cell_date.border = cell_border

        # ── Kolom C: Waktu ────────────────────────────────────────────────
        cell_time = self._worksheet.cell(row=row, column=3)
        cell_time.value = now.strftime("%H:%M:%S")
        cell_time.fill = row_fill
        cell_time.font = Font(size=11, color="2C3E50")
        cell_time.alignment = Alignment(horizontal="center", vertical="center")
        cell_time.border = cell_border

        # ── Kolom D: Screenshot (gambar) ──────────────────────────────────
        cell_img = self._worksheet.cell(row=row, column=4)
        cell_img.fill = row_fill
        cell_img.border = cell_border

        # Atur tinggi baris
        self._worksheet.row_dimensions[row].height = EXCEL_ROW_HEIGHT

        # Sisipkan thumbnail
        self._insert_thumbnail(image, row)

        # Simpan — file temp BELUM dihapus
        self._save()

        logger.info(f"Excel entry #{self._entry_number} added | Row {row}")  # type: ignore[attr-defined]

    @property
    def filepath(self) -> Path:
        return self._filepath

    @property
    def entry_count(self) -> int:
        return self._entry_number

    def finalize(self) -> None:
        """
        Finalisasi: rapikan kolom, tambah summary, save, hapus file temp.
        """
        self._add_summary_row()
        self._apply_column_widths()
        self._save()

        logger.info(f"Excel report finalized: {self._filepath.name}")
        logger.info(f"Total entries this session: {self._entry_number}")

        # Hapus SEMUA file temp setelah save() terakhir
        cleaned = 0
        for tmp in self._temp_files:
            try:
                if tmp.exists():
                    tmp.unlink()
                    cleaned += 1
            except Exception as e:
                logger.warning(f"Gagal hapus temp file {tmp.name}: {e}")
        self._temp_files.clear()
        if cleaned:
            logger.info(f"Cleaned {cleaned} temporary thumbnail files.")

    # =========================================================================
    # PRIVATE — SETUP
    # =========================================================================

    def _create_new(self) -> None:
        """Buat workbook baru dari awal dengan tampilan profesional."""
        self._workbook = Workbook()
        self._worksheet = self._workbook.active
        self._worksheet.title = "Activity Log"
        self._worksheet.sheet_properties.tabColor = COLOR_ACCENT

        # Matikan gridlines — tampilan lebih bersih
        self._worksheet.sheet_view.showGridLines = False

        self._row_counter = 0
        self._entry_number = 0

        self._setup_title_block()
        self._setup_header()
        self._save()
        logger.info(f"Excel report created: {self._filepath.name}")

    def _setup_title_block(self) -> None:
        """
        Buat blok judul di baris 1-2 sebelum header kolom.

        Menggunakan sel biasa (TANPA merge_cells) karena merge_cells
        menyebabkan I/O error saat workbook dibuka ulang untuk append.
        """
        now = datetime.now()

        # Baris 1: Judul utama — hanya di kolom A, tanpa merge
        title_cell = self._worksheet.cell(row=1, column=1)
        title_cell.value = "SCREEN ACTIVITY LOGGER"
        title_cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
        title_cell.font = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self._worksheet.row_dimensions[1].height = 35

        # Baris 1 kolom B-D: sama warna tapi kosong (tidak merge)
        for col in range(2, 5):
            cell = self._worksheet.cell(row=1, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)

        # Baris 2: Info laporan
        info_cell = self._worksheet.cell(row=2, column=1)
        info_cell.value = (
            f"Tanggal: {now.strftime('%d %B %Y')}   |   "
            f"Dibuat: {now.strftime('%H:%M:%S')}   |   "
            f"File: {self._filepath.name}"
        )
        info_cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TITLE_ROW)
        info_cell.font = Font(size=10, color="1A1A2E", italic=True)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        self._worksheet.row_dimensions[2].height = 18

        # Baris 2 kolom B-D: sama warna tapi kosong
        for col in range(2, 5):
            cell = self._worksheet.cell(row=2, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TITLE_ROW)

        # row_counter sekarang di baris 2, header akan di baris 3
        self._row_counter = 2

    def _setup_header(self) -> None:
        """
        Buat baris header kolom (baris 3) dengan style gelap profesional.

        Style:
        - Background: biru sangat gelap
        - Font: putih, bold, size 12
        - Alignment: center
        - Border bawah tebal sebagai pemisah visual
        """
        self._row_counter += 1  # Baris 3 = header
        header_row = self._row_counter

        thick = Side(border_style="medium", color=COLOR_ACCENT)
        thin  = Side(border_style="thin",   color="AAAAAA")

        header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
        header_font = Font(color=COLOR_HEADER_FONT, bold=True, size=12, name="Calibri")
        header_align = Alignment(horizontal="center", vertical="center")

        col_labels = ["No", "Tanggal", "Waktu", "Screenshot"]
        for col_idx, label in enumerate(col_labels, start=1):
            cell = self._worksheet.cell(row=header_row, column=col_idx)
            cell.value = label
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = Border(
                left=thin, right=thin,
                top=thin, bottom=thick  # Garis tebal di bawah header
            )

        self._worksheet.row_dimensions[header_row].height = 28

        # Freeze dari baris setelah header agar title + header selalu terlihat
        self._worksheet.freeze_panes = f"A{header_row + 1}"

    def _add_summary_row(self) -> None:
        """
        Tambahkan baris ringkasan di bawah semua data.
        Tanpa merge_cells untuk menghindari I/O error.
        """
        if self._entry_number == 0:
            return

        self._row_counter += 1
        summary_row = self._row_counter

        summary_text = (
            f"Monitoring selesai  |  "
            f"Total screenshot: {self._entry_number}  |  "
            f"Waktu selesai: {datetime.now().strftime('%H:%M:%S')}"
        )

        summary_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
        summary_font = Font(color="1B5E20", bold=True, size=11)
        summary_align_center = Alignment(horizontal="center", vertical="center")

        thick = Side(border_style="medium", color="4CAF50")
        summary_border = Border(left=thick, right=thick, top=thick, bottom=thick)

        # Kolom A: teks ringkasan
        cell_a = self._worksheet.cell(row=summary_row, column=1)
        cell_a.value = summary_text
        cell_a.fill = summary_fill
        cell_a.font = summary_font
        cell_a.alignment = summary_align_center
        cell_a.border = summary_border

        # Kolom B-D: warna sama, kosong
        for col in range(2, 5):
            cell = self._worksheet.cell(row=summary_row, column=col)
            cell.fill = summary_fill
            cell.border = summary_border

        self._worksheet.row_dimensions[summary_row].height = 24

    # =========================================================================
    # PRIVATE — THUMBNAIL
    # =========================================================================

    def _insert_thumbnail(self, image: PilImage.Image, row: int) -> None:
        """
        Buat thumbnail HD dari PIL Image dan sisipkan ke kolom D.

        Kualitas tinggi dicapai dengan:
        1. Resize menggunakan LANCZOS (algoritma terbaik untuk downscale).
        2. Ukuran 480px lebar — cukup besar untuk membaca teks di layar.
        3. Tinggi dihitung proporsional dari dimensi gambar asli.
        4. Sharpening ringan via Pillow ImageFilter untuk ketajaman teks.

        File temp dikumpulkan di _temp_files dan dihapus di finalize().
        """
        try:
            thumb = image.copy()

            # Konversi ke RGB
            if thumb.mode in ("RGBA", "P", "LA"):
                thumb = thumb.convert("RGB")

            # Hitung tinggi proporsional dari gambar asli
            orig_w, orig_h = thumb.size
            scale   = THUMBNAIL_WIDTH / orig_w
            thumb_h = int(orig_h * scale)

            # Resize dengan LANCZOS — kualitas terbaik untuk downscale
            thumb = thumb.resize((THUMBNAIL_WIDTH, thumb_h), PilImage.LANCZOS)

            # Sharpen ringan agar teks lebih tajam setelah resize
            from PIL import ImageEnhance
            enhancer = ImageEnhance.Sharpness(thumb)
            thumb = enhancer.enhance(1.4)  # 1.0 = normal, 2.0 = sangat tajam

            # Simpan ke file temp dengan kualitas PNG maksimal
            with tempfile.NamedTemporaryFile(
                suffix=".png", delete=False, prefix="sal_thumb_"
            ) as tmp:
                temp_path = Path(tmp.name)
                thumb.save(
                    str(temp_path),
                    format="PNG",
                    optimize=False,  # Jangan kompres — prioritas kualitas
                )

            self._temp_files.append(temp_path)

            # Buat ExcelImage dengan ukuran yang tepat
            excel_img        = ExcelImage(str(temp_path))
            excel_img.width  = THUMBNAIL_WIDTH
            excel_img.height = thumb_h

            # Sisipkan ke kolom D
            self._worksheet.add_image(excel_img, f"D{row}")

            # Lebar kolom D: 1 unit Excel ≈ 7px, tambah sedikit padding
            col_width = int(THUMBNAIL_WIDTH / 7) + 3
            self._worksheet.column_dimensions["D"].width = col_width

            # Tinggi baris dalam points (1pt ≈ 1.33px)
            row_height = max(EXCEL_ROW_HEIGHT, thumb_h / 1.33 + 5)
            self._worksheet.row_dimensions[row].height = row_height

        except Exception as e:
            logger.error(f"Gagal menyisipkan thumbnail baris {row}: {e}")

    # =========================================================================
    # PRIVATE — SAVE & COLUMNS
    # =========================================================================

    def _apply_column_widths(self) -> None:
        """
        Terapkan lebar kolom final.

        Kolom A (No)         :  6 unit
        Kolom B (Tanggal)    : 16 unit
        Kolom C (Waktu)      : 14 unit
        Kolom D (Screenshot) : dihitung dari THUMBNAIL_WIDTH (diset saat insert)
        """
        widths = {"A": 6, "B": 16, "C": 14}
        for col_letter, width in widths.items():
            self._worksheet.column_dimensions[col_letter].width = width

    def _save(self) -> None:
        """
        Simpan workbook. File temp TIDAK dihapus di sini.
        Penghapusan dilakukan di finalize() setelah save() terakhir.
        """
        try:
            self._workbook.save(self._filepath)
        except PermissionError:
            logger.error(
                f"Gagal menyimpan Excel: file sedang dibuka.\n"
                f"Tutup '{self._filepath.name}' lalu coba lagi."
            )
        except Exception as e:
            logger.error(f"Gagal menyimpan Excel: {e}")
